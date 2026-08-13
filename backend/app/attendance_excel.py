"""
attendance_excel.py — Modul BARU Absensi: Export Excel Laporan Absensi
=============================================================================
SATU-SATUNYA laporan di aplikasi ini yang punya export Excel (SELAIN PDF) --
atas permintaan eksplisit spesifikasi Absensi ("Export -> Excel, PDF").
Pemakaian pertama openpyxl di proyek ini (lihat requirements.txt).

Data diambil APA ADANYA lewat attendance_db.get_log_list() -- file ini murni
menyusun lembar Excel, TIDAK menghitung ulang satu angka pun sendiri, sama
seperti prinsip laporan_pdf.py."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import attendance_db

_LABEL_STATUS = {
    "belum_check_in": "Belum Check In", "sedang_bekerja": "Sedang Bekerja",
    "sudah_check_out": "Sudah Check Out", "tidak_check_in": "Tidak Check In",
    "tidak_check_out": "Tidak Check Out",
}
_LABEL_KETEPATAN = {"tepat_waktu": "Tepat Waktu", "terlambat": "Terlambat"}

_HEADER = ["Barber", "Tanggal", "Check In", "Check Out", "Status", "Ketepatan", "Durasi Kerja (menit)",
           "Jarak Check In (meter)", "Jarak Check Out (meter)", "Akurasi GPS Check In (meter)"]


def _jam_dari_iso(nilai):
    if not nilai:
        return ""
    return nilai[11:16] if len(nilai) >= 16 else nilai


def buat_excel_absensi_list(tanggal: str | None, tanggal_dari: str | None, tanggal_sampai: str | None,
                             barber_id: int | None, status: str | None, tenant_id: int) -> bytes:
    data = attendance_db.get_log_list(tenant_id, tanggal=tanggal, tanggal_dari=tanggal_dari,
                                       tanggal_sampai=tanggal_sampai, barber_id=barber_id, status=status)

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Absensi"

    ws.append(_HEADER)
    for sel in ws[1]:
        sel.font = Font(bold=True, color="FFFFFF")
        sel.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        sel.alignment = Alignment(horizontal="center", vertical="center")

    for r in data:
        ws.append([
            r["nama_barber"], r["tanggal"], _jam_dari_iso(r.get("check_in_at")),
            _jam_dari_iso(r.get("check_out_at")), _LABEL_STATUS.get(r["status"], r["status"]),
            _LABEL_KETEPATAN.get(r.get("check_in_status"), ""), r.get("durasi_kerja_menit") or "",
            round(r["check_in_jarak_meter"]) if r.get("check_in_jarak_meter") is not None else "",
            round(r["check_out_jarak_meter"]) if r.get("check_out_jarak_meter") is not None else "",
            round(r["check_in_accuracy"]) if r.get("check_in_accuracy") is not None else "",
        ])

    lebar = [22, 12, 10, 10, 16, 12, 20, 20, 20, 24]
    for i, w in enumerate(lebar, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
